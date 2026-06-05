# -*- coding: utf-8 -*-
"""
XBRL 서비스 레이어 (service.py)

XBRLParser 인스턴스의 생명주기 관리와 비즈니스 로직을 캡슐화합니다.
FastAPI 라우터뿐 아니라 CLI, MCP 등 다양한 프론트엔드에서 동일한
인터페이스로 XBRL 데이터를 활용할 수 있도록 설계되었습니다.

사용 예시 (CLI):
    >>> from backend.service import XBRLService
    >>> svc = XBRLService()
    >>> svc.load_from_zip("path/to/report.zip")
    >>> reports = svc.get_reports()
    >>> data = svc.get_report_data("role_uri_here")
    >>> df = svc.get_report_dataframe("role_uri_here")

사용 예시 (FastAPI):
    >>> svc = XBRLService()
    >>> svc.load_default()  # xbrl sample 폴더 자동 탐색
    >>> # 라우터에서 svc.get_reports() 등 호출
"""

import io
import os
import re
import math
import logging

import pandas as pd

from backend.xbrl_parser import XBRLParser

# ── 로거 설정 ──
logger = logging.getLogger(__name__)


def clean_json_value(v):
    """JSON 직렬화에 안전하지 않은 값(NaN, Inf, pd.NA)을 None으로 변환합니다.

    Pandas/NumPy의 float NaN이나 Inf는 JSON 표준에서 허용되지 않으므로,
    API 응답을 생성하기 전에 이 함수로 정제해야 합니다.

    Args:
        v: 검사할 값

    Returns:
        정제된 값 (문제가 있으면 None)
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _clean_record(record: dict) -> dict:
    """딕셔너리 내 모든 값을 JSON-safe하게 정제합니다."""
    return {k: clean_json_value(v) for k, v in record.items()}


class XBRLService:
    """XBRL 뷰어의 핵심 비즈니스 로직을 제공하는 서비스 클래스.

    파서 인스턴스 관리, 보고서 조회, 내보내기, 수식 분석 등의 기능을
    프론트엔드 독립적으로 제공합니다.

    Attributes:
        parser (XBRLParser | None): 현재 활성화된 파서 인스턴스
        file_name (str): 현재 로드된 파일명
    """

    # 기본 XBRL 샘플 디렉토리 (프로젝트 루트 기준)
    DEFAULT_SAMPLE_DIR = os.path.abspath(
        os.path.join(os.path.dirname(__file__), '..', 'xbrl sample')
    )

    def __init__(self):
        """서비스를 초기화합니다. 파서는 아직 로드되지 않은 상태입니다."""
        self.parser: XBRLParser | None = None
        self.file_name: str = ""

    # ──────────────────────────────────────────────────────────────────────
    # 파서 로드 메서드
    # ──────────────────────────────────────────────────────────────────────

    def load_from_path(self, path: str, name: str = ""):
        """디렉토리 또는 ZIP 파일 경로로부터 XBRL 파서를 로드합니다.

        Args:
            path: XBRL 디렉토리 또는 ZIP 파일 경로
            name: 표시용 파일명 (빈 문자열이면 경로에서 자동 추출)

        Raises:
            ValueError: 경로가 유효하지 않은 경우
            FileNotFoundError: 경로가 존재하지 않는 경우
        """
        if not os.path.exists(path):
            raise FileNotFoundError(f"경로를 찾을 수 없습니다: {path}")

        # 기존 파서가 있으면 정리
        if self.parser:
            self.parser.close()

        self.parser = XBRLParser(path)
        self.file_name = name or os.path.basename(path)
        logger.info(f"XBRL 파일 로드 완료: {self.file_name}")

    def load_from_zip(self, zip_path: str, name: str = ""):
        """ZIP 파일 경로로부터 XBRL 파서를 로드합니다.

        load_from_path의 편의 메서드입니다.

        Args:
            zip_path: ZIP 파일 경로
            name: 표시용 파일명
        """
        self.load_from_path(zip_path, name or os.path.basename(zip_path))

    def load_default(self) -> bool:
        """기본 xbrl sample 폴더에서 첫 번째 ZIP 파일을 자동으로 로드합니다.

        Returns:
            로드 성공 여부
        """
        sample_dir = self.DEFAULT_SAMPLE_DIR
        if not os.path.isdir(sample_dir):
            logger.warning(f"기본 샘플 디렉토리를 찾을 수 없습니다: {sample_dir}")
            return False

        # ZIP 파일 중 첫 번째를 자동 로드
        zip_files = sorted([
            f for f in os.listdir(sample_dir) if f.endswith('.zip')
        ])
        if not zip_files:
            logger.warning(f"샘플 디렉토리에 ZIP 파일이 없습니다: {sample_dir}")
            return False

        zip_path = os.path.join(sample_dir, zip_files[0])
        try:
            self.load_from_path(zip_path, zip_files[0])
            return True
        except Exception as e:
            logger.error(f"기본 파일 로드 실패: {e}")
            return False

    @property
    def is_loaded(self) -> bool:
        """파서가 로드되어 있는지 확인합니다."""
        return self.parser is not None

    # ──────────────────────────────────────────────────────────────────────
    # 보고서 조회 메서드
    # ──────────────────────────────────────────────────────────────────────

    def get_status(self) -> dict:
        """현재 서비스 상태를 딕셔너리로 반환합니다.

        Returns:
            상태 정보 딕셔너리
        """
        if not self.is_loaded:
            return {
                "status": "ready",
                "file_loaded": False,
                "message": "XBRL 파일이 로드되지 않았습니다. ZIP 파일을 업로드해 주세요.",
            }
        return {
            "status": "ready",
            "file_loaded": True,
            "file_name": self.file_name,
            "instants": self.parser.instants,
            "durations": self.parser.durations,
        }

    def get_reports(self) -> list:
        """보고서 시트 목록을 반환합니다.

        Returns:
            시트 정보 딕셔너리 리스트

        Raises:
            RuntimeError: 파서가 로드되지 않은 경우
        """
        self._ensure_loaded()
        return self.parser.get_sheet_list()

    def get_report_data(self, role_uri: str, active_axis: str = None) -> dict:
        """특정 역할의 보고서 데이터를 JSON-safe한 형태로 반환합니다.

        Args:
            role_uri: 조회할 역할 URI
            active_axis: 선택된 활성 축의 concept_id

        Returns:
            {period_type, periods, data, is_equity_statement, equity_columns} 딕셔너리
        """
        self._ensure_loaded()

        df, period_type, periods = self.parser.get_dataframe_for_role(role_uri, active_axis=active_axis)

        if df.empty:
            return {"data": [], "period_type": period_type, "periods": []}

        # 기간 문자열 변환
        periods_str = []
        for p in periods:
            if isinstance(p, tuple):
                periods_str.append(f"{p[0]} ~ {p[1]}")
            else:
                periods_str.append(str(p))

        # DataFrame → dict 변환 후 NaN/Inf 정제
        records = df.to_dict(orient="records")
        cleaned_records = [_clean_record(r) for r in records]

        res = {
            "period_type": period_type,
            "periods": periods_str,
            "data": cleaned_records,
        }
        if df.attrs.get('is_equity'):
            res["is_equity_statement"] = True
            res["equity_columns"] = df.attrs.get('equity_columns')
        elif df.attrs.get('is_dimensional'):
            res["is_dimensional"] = True
            res["dimensional_columns"] = df.attrs.get('dimensional_columns')
            res["dimensional_axes"] = df.attrs.get('dimensional_axes')
        return res

    def get_report_dataframe(self, role_uri: str) -> pd.DataFrame:
        """특정 역할의 보고서 데이터를 Pandas DataFrame으로 반환합니다.

        CLI에서 직접 DataFrame을 다루고 싶을 때 사용합니다.

        Args:
            role_uri: 조회할 역할 URI

        Returns:
            계층 구조가 반영된 DataFrame
        """
        self._ensure_loaded()
        df, _period_type, _periods = self.parser.get_dataframe_for_role(role_uri)
        return df

    # ──────────────────────────────────────────────────────────────────────
    # 내보내기 메서드
    # ──────────────────────────────────────────────────────────────────────

    def export_report(self, role_uri: str, fmt: str = "excel", active_axis: str = None) -> tuple:
        """보고서를 Excel 또는 CSV 형식으로 내보냅니다.

        Args:
            role_uri: 내보낼 역할 URI
            fmt: 'excel' 또는 'csv'
            active_axis: 선택된 활성 축의 concept_id

        Returns:
            (BytesIO 또는 StringIO, media_type, filename) 튜플

        Raises:
            RuntimeError: 파서가 로드되지 않은 경우
            ValueError: 데이터가 없는 경우
        """
        self._ensure_loaded()

        df, period_type, periods = self.parser.get_dataframe_for_role(role_uri, active_axis=active_axis)
        if df.empty:
            raise ValueError("해당 시트에 데이터가 없습니다.")

        # 시트 제목 생성
        sheet_title = self.parser.role_definitions.get(role_uri, "Report")
        # '[D210000] 재무상태표' → '재무상태표'
        if ']' in sheet_title:
            sheet_title = sheet_title.split(']')[-1].strip()
        safe_title = "".join(
            c for c in sheet_title if c.isalnum() or c in " _-"
        )[:30]

        # 기간 문자열 변환
        periods_str = []
        for p in periods:
            if isinstance(p, tuple):
                periods_str.append(f"{p[0]} ~ {p[1]}")
            else:
                periods_str.append(str(p))

        # 내보내기용 컬럼 구성
        if df.attrs.get('is_equity'):
            display_cols = ['label_ko', 'concept_id', 'depth']
            col_names = ['항목명 (Korean)', 'Concept ID', 'Depth']
            for col_meta in df.attrs.get('equity_columns'):
                display_cols.append(col_meta['key'])
                period_label = '당기' if col_meta['period'] == 't' else ('전기' if col_meta['period'] == 't1' else '전전기')
                col_names.append(f"{col_meta['label']} ({period_label} {col_meta['period_str']})")
        elif df.attrs.get('is_dimensional'):
            display_cols = ['label_ko', 'concept_id', 'depth']
            col_names = ['항목명 (Korean)', 'Concept ID', 'Depth']
            for col_meta in df.attrs.get('dimensional_columns'):
                display_cols.append(col_meta['key'])
                period_label = '당기' if col_meta['period'] == 't' else ('전기' if col_meta['period'] == 't1' else '전전기')
                col_names.append(f"{col_meta['label']} ({period_label} {col_meta['period_str']})")
        else:
            display_cols = ['label_ko', 'concept_id', 'depth']
            col_names = ['항목명 (Korean)', 'Concept ID', 'Depth']

            period_labels = ['당기', '전기', '전전기']
            for i, p_str in enumerate(periods_str):
                suffix = f" ({period_labels[i]} {p_str})" if i < len(period_labels) else f" ({p_str})"
                t_key = ['t', 't1', 't2'][i]
                display_cols.append(f"consolidated_{t_key}")
                col_names.append(f"연결{suffix}")
                display_cols.append(f"separate_{t_key}")
                col_names.append(f"별도{suffix}")

        df_export = df[display_cols].copy()
        df_export.columns = col_names

        if fmt.lower() == "csv":
            return self._export_csv(df_export, safe_title)
        else:
            return self._export_excel(df_export, df, safe_title, col_names)

    def _export_csv(self, df_export: pd.DataFrame, title: str) -> tuple:
        """CSV 형식으로 내보냅니다."""
        stream = io.StringIO()
        df_export.to_csv(stream, index=False, encoding='utf-8-sig')
        content = stream.getvalue().encode('utf-8-sig')
        return (
            io.BytesIO(content),
            "text/csv",
            f"{title}.csv",
        )

    def _export_excel(
        self, df_export: pd.DataFrame, df_raw: pd.DataFrame,
        title: str, col_names: list,
    ) -> tuple:
        """서식이 적용된 Excel 파일로 내보냅니다."""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name=title, index=False)

            workbook = writer.book
            worksheet = writer.sheets[title]

            # ── 스타일 정의 ──
            header_fill = PatternFill(
                start_color="1F2937", end_color="1F2937", fill_type="solid"
            )
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            bold_font = Font(name="Calibri", size=10, bold=True)
            normal_font = Font(name="Calibri", size=10)
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB'),
            )

            # ── 헤더 행 서식 ──
            for col_idx in range(1, len(col_names) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            worksheet.row_dimensions[1].height = 28

            # ── 데이터 행 서식 ──
            for row_idx, row in df_raw.iterrows():
                excel_row = row_idx + 2
                depth = int(row['depth'])
                is_abstract = bool(row['is_abstract'])

                worksheet.row_dimensions[excel_row].height = 20

                # 깊이에 따른 들여쓰기
                label_cell = worksheet.cell(row=excel_row, column=1)
                indent_spaces = "    " * depth
                label_cell.value = f"{indent_spaces}{row['label_ko']}"

                # 추상 행은 볼드 처리
                row_font = bold_font if is_abstract else normal_font

                for col_idx in range(1, len(col_names) + 1):
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.font = row_font
                    cell.border = thin_border

                    if col_idx > 3:  # 숫자 컬럼
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )
                        if cell.value is not None:
                            cell.number_format = '#,##0;[Red]-#,##0;"-"'
                    elif col_idx == 3:  # Depth 컬럼
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:  # 텍스트 컬럼
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center"
                        )

            # ── 열 너비 자동 조정 ──
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # 항목명 컬럼은 고정 너비
            worksheet.column_dimensions['A'].width = 45

        output.seek(0)
        return (
            output,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{title}.xlsx",
        )

    # ──────────────────────────────────────────────────────────────────────
    # 재무 분석 (수식 연산)
    # ──────────────────────────────────────────────────────────────────────

    def analyze(self, formulas: dict) -> dict:
        """사용자 정의 수식을 XBRL 팩트 값으로 평가합니다.

        수식의 변수(Concept ID 형태)를 실제 팩트 값으로 치환하여 계산합니다.
        당기/전기에 대해 연결/별도 각각의 결과를 반환합니다.

        Args:
            formulas: {이름: 수식 문자열} 딕셔너리
                예: {"current_ratio": "ifrs-full_CurrentAssets / ifrs-full_CurrentLiabilities"}

        Returns:
            {이름: {formula, consolidated_t, consolidated_t1,
                    separate_t, separate_t1, values_used}} 딕셔너리
        """
        self._ensure_loaded()
        parser = self.parser
        results = {}

        for name, expr in formulas.items():
            expr = expr.strip()

            # 수식에서 Concept ID 토큰 추출 (예: ifrs-full_CurrentAssets)
            tokens = re.findall(r'[a-zA-Z0-9_\-]+_[a-zA-Z0-9_\-]+', expr)

            results[name] = {
                "formula": expr,
                "consolidated_t": None,
                "consolidated_t1": None,
                "separate_t": None,
                "separate_t1": None,
                "values_used": {},
            }

            # ── 시점형(instant) 기간으로 먼저 평가 ──
            instant_configs = [
                ("consolidated_t", parser.instants[0] if parser.instants else None, False, 'instant'),
                ("consolidated_t1", parser.instants[1] if len(parser.instants) > 1 else None, False, 'instant'),
                ("separate_t", parser.instants[0] if parser.instants else None, True, 'instant'),
                ("separate_t1", parser.instants[1] if len(parser.instants) > 1 else None, True, 'instant'),
            ]

            for cfg_name, actual_p, is_sep, p_type in instant_configs:
                val, local_env = self._evaluate_formula(
                    expr, tokens, actual_p, is_sep, p_type
                )
                results[name][cfg_name] = clean_json_value(val)
                self._record_audit_trail(results[name], cfg_name, local_env)

            # ── 시점형 결과가 모두 None이면 기간형(duration)으로 재시도 ──
            all_none = all(
                results[name][c] is None
                for c in ["consolidated_t", "consolidated_t1", "separate_t", "separate_t1"]
            )
            if all_none:
                dur_configs = [
                    ("consolidated_t", parser.durations[0] if parser.durations else None, False, 'duration'),
                    ("consolidated_t1", parser.durations[1] if len(parser.durations) > 1 else None, False, 'duration'),
                    ("separate_t", parser.durations[0] if parser.durations else None, True, 'duration'),
                    ("separate_t1", parser.durations[1] if len(parser.durations) > 1 else None, True, 'duration'),
                ]
                for cfg_name, actual_p, is_sep, p_type in dur_configs:
                    val, local_env = self._evaluate_formula(
                        expr, tokens, actual_p, is_sep, p_type
                    )
                    results[name][cfg_name] = clean_json_value(val)
                    self._record_audit_trail(results[name], cfg_name, local_env)

        return results

    def _evaluate_formula(
        self, expr: str, tokens: list, period, is_separate: bool, period_type: str,
    ) -> tuple:
        """수식의 변수를 팩트 값으로 치환하고 평가합니다.

        시점형/기간형 교차 폴백을 지원합니다: 해당 period_type에서 값을 찾지 못하면
        반대 유형에서도 탐색합니다.

        Args:
            expr: 수식 문자열
            tokens: 수식에서 추출한 Concept ID 목록
            period: 대상 보고 기간
            is_separate: 별도 재무제표 여부
            period_type: 'instant' 또는 'duration'

        Returns:
            (평가 결과, 사용된 변수 환경) 튜플
        """
        parser = self.parser
        if not period:
            return None, {}

        local_env = {}
        for t in tokens:
            # 1차: 지정된 period_type으로 조회
            v = parser.get_fact_value(
                t, period, is_separate=is_separate, period_type=period_type
            )
            # 2차: 반대 유형으로 폴백 시도
            if v is None:
                other_type = 'duration' if period_type == 'instant' else 'instant'
                if other_type == 'duration':
                    # instant 날짜에 매칭되는 duration 기간 탐색
                    matching_dur = next(
                        (dur for dur in parser.durations if dur[1] == period),
                        None,
                    )
                    if matching_dur:
                        v = parser.get_fact_value(
                            t, matching_dur, is_separate=is_separate, period_type='duration'
                        )
                else:
                    # duration 종료일에 매칭되는 instant 날짜 탐색
                    target_inst = period[1]  # 종료일
                    v = parser.get_fact_value(
                        t, target_inst, is_separate=is_separate, period_type='instant'
                    )

            if v is not None:
                local_env[t] = float(v) if isinstance(v, (int, float)) else 0.0

        # ── 수식 평가: 변수를 실제 값으로 치환 ──
        # 긴 토큰부터 치환하여 부분 매칭 방지
        sorted_tokens = sorted(tokens, key=len, reverse=True)
        eval_expr = expr
        for t in sorted_tokens:
            if t in local_env:
                eval_expr = eval_expr.replace(t, str(local_env[t]))
            else:
                eval_expr = eval_expr.replace(t, "0.0")

        # 안전한 문자만 허용 (숫자, 연산자, 괄호, 공백, 소수점)
        clean_expr = re.sub(r'[^0-9\+\-\*\/\(\)\s\.]', '', eval_expr)
        if not clean_expr.strip():
            return None, {}

        try:
            val = eval(clean_expr)  # noqa: S307 — 입력이 sanitize됨
            return val, local_env
        except Exception as e:
            logger.warning(f"수식 평가 실패 ({clean_expr}): {e}")
            return None, {}

    def _record_audit_trail(self, result_entry: dict, cfg_name: str, local_env: dict):
        """수식 평가에 사용된 변수값을 감사 추적(audit trail)에 기록합니다."""
        parser = self.parser
        for t, val_used in local_env.items():
            if t not in result_entry["values_used"]:
                result_entry["values_used"][t] = {}
            lbl = parser.labels_ko.get(t, t)
            result_entry["values_used"][t]["label"] = lbl
            result_entry["values_used"][t][cfg_name] = clean_json_value(val_used)

    # ──────────────────────────────────────────────────────────────────────
    # 내부 유틸리티
    # ──────────────────────────────────────────────────────────────────────

    def _ensure_loaded(self):
        """파서가 로드되어 있는지 확인하고, 없으면 예외를 발생시킵니다."""
        if not self.is_loaded:
            raise RuntimeError("XBRL 파일이 로드되지 않았습니다.")
